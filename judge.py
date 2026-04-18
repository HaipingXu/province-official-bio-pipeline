"""
Judge module: LLM-based arbitration for disputed fields across pipeline steps.

Provides judge_step1(), judge_step2(), judge_step3() entry points
and Excel styling helpers.
"""

import json
import logging
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from config import (
    LOGS_DIR, OUTPUT_DIR,
    DEFAULT_WORKERS,
    EP_CHECK_FIELDS, EPISODE_FIELDS,
)
from utils import (
    extract_json, load_prompt, llm_chat,
    RoundRobinClientPool, LLMConfig, to_float_date,
    load_json_cache, save_json_cache,
)

logger = logging.getLogger(__name__)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ── Judge reference prompts (loaded once at import) ──────────────────────────

def _load_judge_references() -> str:
    parts: list[str] = []
    for prompt_name, title in [
        ("step1_extraction", "Step 1 字段提取规则（供职单位/职务/组织标签/标志位/任职地）"),
        ("step2_rank",       "Step 2 行政级别判断规则（正国级→副科级，10级体系）"),
        ("step3_labeling",   "Step 3 标签规则（升迁_省长/升迁_省委书记/本省提拔/本省学习/落马）"),
    ]:
        try:
            content = load_prompt(prompt_name)
            parts.append(f"\n\n---\n\n## 裁判参考：{title}\n\n{content}")
        except Exception as e:
            logger.warning(f"[judge] 无法加载参考 prompt {prompt_name}: {e}")
    return "".join(parts)


_JUDGE_REFERENCE: str | None = None


def _get_judge_reference() -> str:
    """Lazy-load judge reference prompts on first access."""
    global _JUDGE_REFERENCE
    if _JUDGE_REFERENCE is None:
        _JUDGE_REFERENCE = _load_judge_references()
    return _JUDGE_REFERENCE


# ── Judge preambles ──────────────────────────────────────────────────────────

_JUDGE_PREAMBLE = (
    "你是一名中国政治数据核查专家，根据原文对两个LLM提取结果做裁判。\n\n"
    "核心规则：\n"
    "1. 学习经历（本科、研究生、博士、进修、培训等）也是条目。\n"
    "2. 如果两方都不完全正确，你可以给出自己根据原文判断的正确值（verdict=自行修正）。\n"
    "3. 党委系统供职单位必须使用全称带「中共」前缀（如「中共深圳市委」而非「深圳市委」）。\n"
    "4. 除党委命名规则外，其他字段忠实于百度百科原文用词。\n"
    "5. 【标志位字段约束】当裁判「标志位」字段时，correct_value 必须且只能是以下24个合法值之一，"
    "禁止使用「|」拼接、禁止添加括注或诊断说明：\n"
    "   市委书记、市长、市委副书记（市长）、市委副书记（非市长）、副市长（常委）、副市长（非常委）、"
    "市常委（其他）、市组织部长、省委书记、省长、省委副书记（省长）、省委副书记（非省长）、"
    "副省长（常委）、副省长（非常委）、省常委（其他）、省组织部长、省组织部副部长、"
    "军队、学习进修、高校/党校任职、秘书、政治局常委、政治局委员、无\n"
    "   若当前条目同时含多个职务，选择本条目主职对应的标志位，不得合并。\n\n"
    "输出JSON格式：\n"
    "{\"verdict\": \"采纳LLM1\"|\"采纳LLM2\"|\"自行修正\"|\"两者均存疑\", "
    "\"correct_value\": \"仅当verdict=自行修正时填写\", "
    "\"confidence\": 0-100, "
    "\"reason\": \"<50字理由>\"}\n"
    "confidence 表示你对该裁决的信心程度（0=完全不确定，100=完全确定）。\n"
)

_JUDGE_SYSTEM_BASE = (
    _JUDGE_PREAMBLE +
    "\n如果一次裁判多个字段，输出JSON对象，key为字段名，value为上述格式：\n"
    "{\"字段A\": {\"verdict\": ..., \"confidence\": 85, \"reason\": ...}, \"字段B\": {...}}\n\n"
    "只输出JSON，无任何其他文字。"
)

_JUDGE_SYSTEM_LABEL_BASE = (
    _JUDGE_PREAMBLE +
    "不要输出任何其他文字、解释或代码块标记。"
)

_JUDGE_SYSTEM_RANK_BASE = (
    "你是一名中国政治数据核查专家，判断行政级别。\n\n"
    "两个LLM对同一职务判断了不同的行政级别，请根据职务和单位判断哪个更准确。\n\n"
    "输出JSON格式：\n"
    "{\"verdict\": \"采纳LLM1\"|\"采纳LLM2\"|\"自行修正\"|\"两者均存疑\", "
    "\"correct_value\": \"仅当verdict=自行修正时填写\", "
    "\"confidence\": 0-100, "
    "\"reason\": \"<50字理由>\"}\n"
    "只输出JSON，无任何其他文字。"
)


def _judge_system() -> str:
    return _JUDGE_SYSTEM_BASE + _get_judge_reference()


def _judge_system_label() -> str:
    return _JUDGE_SYSTEM_LABEL_BASE + _get_judge_reference()


def _judge_system_rank() -> str:
    return _JUDGE_SYSTEM_RANK_BASE + _get_judge_reference()


# ── Cache key builders ────────────────────────────────────────────────────────

def _ep_cache_key(name: str, row: dict) -> str:
    return (
        f"{name}||ep_batch"
        f"||sl{row.get('source_line', '')}"
        f"||{row.get('LLM1_供职单位', '')}"
        f"||{row.get('LLM1_职务', '')}"
        f"||{row.get('LLM1_起始时间', '')}"
    )


def _sl_group_cache_key(name: str, line_num: int) -> str:
    return f"{name}||sl_group||{line_num}"


def _label_cache_key(name: str, field: str) -> str:
    return f"{name}||label||{field}"


def _rank_cache_key(name: str, episode_idx: int) -> str:
    return f"{name}||rank||{episode_idx}"


# ── Core judge call (uses llm_chat) ──────────────────────────────────────────

def _call_judge(system: str, prompt: str, pool: RoundRobinClientPool | None = None, model: str = "") -> dict:
    if not pool:
        return {"verdict": "两者均存疑", "reason": "裁判调用失败: 未提供 pool", "judge_model": "error"}
    try:
        raw = llm_chat(
            pool.next_client(), model,
            system=system, user=prompt,
            temperature=0.0, max_retries=2, seed=None,
        )
        result = extract_json(raw)
        result["judge_model"] = model
        return result
    except Exception as e:
        if "Content Exists Risk" in str(e):
            return {"verdict": "两者均存疑", "reason": "内容安全拦截", "judge_model": "blocked"}
        return {"verdict": "两者均存疑", "reason": f"裁判调用失败: {e}", "judge_model": "error"}


# ── Judge helpers ─────────────────────────────────────────────────────────────

def judge_source_line_group(
    name: str, line_num: int, raw_text: str,
    ds_episodes: list[dict], vf_episodes: list[dict],
    pool: RoundRobinClientPool | None = None, model: str = "",
) -> dict:
    all_fields = EPISODE_FIELDS

    def _fmt(eps):
        lines = []
        for i, ep in enumerate(eps, 1):
            parts = [f"{f}={ep.get(f, '')}" for f in all_fields]
            lines.append(f"  #{i}: " + ", ".join(parts))
        return "\n".join(lines) if lines else "  （无）"

    prompt = (
        f"官员：{name}\n"
        f"原文行 L{line_num:02d}: {raw_text}\n\n"
        f"LLM1从此行提取了 {len(ds_episodes)} 条经历：\n{_fmt(ds_episodes)}\n\n"
        f"LLM2从此行提取了 {len(vf_episodes)} 条经历：\n{_fmt(vf_episodes)}\n\n"
        "两个LLM对这一行原文的拆分方式不同。请判断哪种更准确，并输出最终正确版本。\n\n"
        "判断规则：\n"
        "- 同一时间段在同一单位的多个职务 → 合并为一条（职务用顿号连接）\n"
        "- 不同时间段或不同单位 → 拆分为多条\n"
        "- 党组书记/党组成员与行政职务同期同单位 → 合并\n\n"
        "同时检查所有字段的准确性：\n"
        "- 组织标签：是否与供职单位匹配（如'市人民政府'应标'地级市政府'）\n"
        "- 供职单位：是否含完整层级（如学院名应含大学名前缀）\n"
        "- 中央/地方：中央部委/直属机构/全国人大政协='中央'，省市县='地方'\n"
        "- 任职地：省份用全称（如'广东省'非'广东'），直辖市省市均填\n\n"
        f"输出JSON:\n"
        f"{{\n"
        f"  \"adopt\": \"LLM1\" 或 \"LLM2\",\n"
        f"  \"confidence\": 0-100,\n"
        f"  \"reason\": \"<50字理由>\",\n"
        f"  \"episodes\": [\n"
        f"    {{\n"
        f"      \"source_line\": {line_num},\n"
        f"      \"起始时间\": \"YYYY.MM\",\n"
        f"      \"终止时间\": \"YYYY.MM\",\n"
        f"      \"组织标签\": \"...\",\n"
        f"      \"供职单位\": \"...\",\n"
        f"      \"职务\": \"...\",\n"
        f"      \"任职地（省）\": \"...\",\n"
        f"      \"任职地（市）\": \"...\",\n"
        f"      \"中央/地方\": \"中央或地方\"\n"
        f"    }}\n"
        f"  ]\n"
        f"}}\n"
        f"confidence 表示你对该裁决的信心程度。\n"
        f"episodes数组包含最终正确的经历条目。\n"
        f"只输出JSON，无任何其他文字。"
    )
    return _call_judge(_judge_system(), prompt, pool=pool, model=model)


def judge_episode_batch(
    name: str, disputed_fields: list[str],
    llm1_values: dict[str, str], llm2_values: dict[str, str],
    ref_llm1: str, ref_llm2: str,
    pool: RoundRobinClientPool | None = None, model: str = "",
) -> dict[str, dict]:
    if not disputed_fields:
        return {}

    field_lines = []
    for f in disputed_fields:
        field_lines.append(f"  {f}: LLM1=「{llm1_values.get(f, '')}」 LLM2=「{llm2_values.get(f, '')}」")
    fields_block = "\n".join(field_lines)

    prompt = (
        f"官员：{name}\n\n"
        f"LLM1来源行：{ref_llm1 or '（无）'}\n"
        f"LLM2来源行：{ref_llm2 or '（无）'}\n\n"
        f"以下 {len(disputed_fields)} 个字段存在争议：\n{fields_block}\n\n"
        "请对每个字段分别裁判。\n"
        "输出JSON对象，key为字段名，value含verdict/reason/correct_value。"
    )

    result = _call_judge(_judge_system(), prompt, pool=pool, model=model)

    parsed: dict[str, dict] = {}
    if len(disputed_fields) == 1:
        f = disputed_fields[0]
        if "verdict" in result:
            parsed[f] = result
        elif f in result:
            parsed[f] = result[f]
        else:
            parsed[f] = {"verdict": "两者均存疑", "reason": "解析失败"}
    else:
        for f in disputed_fields:
            if f in result and isinstance(result[f], dict):
                parsed[f] = result[f]
            elif "verdict" in result:
                parsed[f] = result
            else:
                parsed[f] = {"verdict": "两者均存疑", "reason": "批量裁判未覆盖此字段"}

    model_tag = result.get("judge_model", model)
    for f in parsed:
        parsed[f]["judge_model"] = model_tag

    return parsed


def get_judge_decision(
    name: str, field: str, scope: str,
    llm1_value: str, llm1_reason: str,
    llm2_value: str, llm2_reason: str,
    original_text_snippet: str,
    pool: RoundRobinClientPool | None = None, model: str = "",
) -> dict:
    prompt = (
        f"官员：{name}  字段：{field}（{scope}）\n\n"
        f"原文依据：{original_text_snippet or '（无）'}\n\n"
        f"LLM1提取：{llm1_value}\n"
        f"LLM1依据：{llm1_reason or '（无）'}\n\n"
        f"LLM2提取：{llm2_value}\n"
        f"LLM2依据：{llm2_reason or '（无）'}\n\n"
        "根据原文和两方依据，判断哪个更准确。如果两方都不正确，请自行给出正确值。"
    )
    return _call_judge(_judge_system_label(), prompt, pool=pool, model=model)


# ── Common concurrent judge executor ─────────────────────────────────────────


def _run_judge_tasks(
    pending: list[tuple[str, dict]],
    judge_fn: Callable[[str, dict], tuple[str, dict]],
    cache: dict,
    max_workers: int,
    step_label: str = "judge",
    model: str = "",
) -> None:
    """Execute judge tasks concurrently, writing results into *cache* in-place."""
    if not pending:
        return

    if max_workers <= 1:
        for ck, kw in pending:
            try:
                _, decision = judge_fn(ck, kw)
                cache[ck] = decision
            except Exception as e:
                logger.error(f"[{step_label} error] {ck}: {e}")
                cache[ck] = {"verdict": "两者均存疑", "reason": f"异常: {e}", "judge_model": model}
    else:
        lock = threading.Lock()
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = {pool.submit(judge_fn, ck, kw): ck for ck, kw in pending}
            for fut in as_completed(futures):
                try:
                    ck, decision = fut.result()
                    with lock:
                        cache[ck] = decision
                except Exception as e:
                    ck = futures[fut]
                    logger.error(f"[{step_label} error] {ck}: {e}")
                    with lock:
                        cache[ck] = {"verdict": "两者均存疑", "reason": f"异常: {e}", "judge_model": model}


# ── Judge Step 1: episode fields + sl_group → merged_episodes.json ──────────

def judge_step1(
    logs_dir: Path,
    officials_dir: Path | None = None,
    force: bool = False,
    max_workers: int = DEFAULT_WORKERS,
    pool: RoundRobinClientPool | None = None,
    model: str = "",
) -> Path:
    """Judge step1 disputes and produce merged_episodes.json."""
    from diff import group_by_source_line
    from text_preprocessor import preprocess_official
    from merged_builder import build_merged_episodes

    logger.info("=== Step1 Judge: episode 字段 + sl_group 裁判 ===")

    diff_path = logs_dir / "step1_diff_report.json"
    diff_report = json.loads(diff_path.read_text(encoding="utf-8"))

    # Build career_lines_by_name for raw text context
    career_lines_by_name: dict[str, dict[int, str]] = {}
    for person in diff_report:
        oname = person.get("official_name", "")
        if oname:
            preprocessed = preprocess_official(oname, officials_dir=officials_dir)
            if preprocessed and preprocessed.get("career_lines"):
                career_lines_by_name[oname] = {
                    cl["line_num"]: cl["raw_text"] for cl in preprocessed["career_lines"]
                }

    # Load judge cache
    judge_cache_path = logs_dir / "step1_judge_decisions.json"
    judge_cache: dict = {}
    if judge_cache_path.exists() and not force:
        try:
            judge_cache = json.loads(judge_cache_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    # Collect disputes
    sl_mismatch_data: dict[tuple[str, int], dict] = {}
    pending_episode_calls: list[tuple[str, dict]] = []
    pending_group_calls: list[tuple[str, dict]] = []

    for person in diff_report:
        name = person["official_name"]
        ds_s1 = person.get("llm1_step1", {})
        vf_s1 = person.get("llm2_step1", {})
        eps_ds = ds_s1.get("episodes", [])
        eps_vf = vf_s1.get("episodes", [])
        cl_map = career_lines_by_name.get(name, {})

        ds_groups = group_by_source_line(eps_ds)
        vf_groups = group_by_source_line(eps_vf)
        all_lines = sorted(set(ds_groups) | set(vf_groups))

        for line_num in all_lines:
            ds_list = ds_groups.get(line_num, [])
            vf_list = vf_groups.get(line_num, [])

            if len(ds_list) != len(vf_list):
                sl_mismatch_data[(name, line_num)] = {
                    "ds_episodes": ds_list, "vf_episodes": vf_list,
                    "raw_text": cl_map.get(line_num, ""),
                }
                cache_key = _sl_group_cache_key(name, line_num)
                if cache_key not in judge_cache:
                    pending_group_calls.append((cache_key, {
                        "name": name, "line_num": line_num,
                        "raw_text": cl_map.get(line_num, ""),
                        "ds_episodes": ds_list, "vf_episodes": vf_list,
                    }))
                continue

            ds_sorted = sorted(ds_list, key=lambda e: e.get("供职单位", ""))
            vf_sorted = sorted(vf_list, key=lambda e: e.get("供职单位", ""))

            for i in range(min(len(ds_sorted), len(vf_sorted))):
                ep_ds = ds_sorted[i]
                ep_vf = vf_sorted[i]
                disputed_fields = []
                row = {"source_line": line_num}
                for f in EP_CHECK_FIELDS:
                    if f == "行政级别":
                        continue  # rank is judged in step2
                    v_ds = str(ep_ds.get(f, ""))
                    v_vf = str(ep_vf.get(f, ""))
                    row[f"LLM1_{f}"] = v_ds
                    row[f"LLM2_{f}"] = v_vf
                    if v_ds != v_vf:
                        disputed_fields.append(f)

                if not disputed_fields:
                    continue

                row["LLM1_供职单位"] = ep_ds.get("供职单位", "")
                row["LLM1_职务"] = ep_ds.get("职务", "")
                row["LLM1_起始时间"] = ep_ds.get("起始时间", "")
                ep_key = _ep_cache_key(name, row)

                all_cached = all(f"{ep_key}||{f}" in judge_cache for f in disputed_fields)
                if not all_cached:
                    llm1_vals = {f: str(row.get(f"LLM1_{f}", "")) for f in disputed_fields}
                    llm2_vals = {f: str(row.get(f"LLM2_{f}", "")) for f in disputed_fields}
                    sl_num = ep_ds.get("source_line", line_num)
                    raw = cl_map.get(sl_num, "")
                    pending_episode_calls.append((ep_key, {
                        "name": name,
                        "disputed_fields": disputed_fields,
                        "llm1_values": llm1_vals,
                        "llm2_values": llm2_vals,
                        "ref_llm1": f"L{sl_num:02d}: {raw}" if raw else f"L{sl_num:02d}",
                        "ref_llm2": "",
                    }))

    total_calls = len(pending_episode_calls) + len(pending_group_calls)

    if total_calls > 0:
        n_fields = sum(len(kw["disputed_fields"]) for _, kw in pending_episode_calls)
        logger.info(f"  裁判调用: {len(pending_episode_calls)} 条履历({n_fields}个字段) + "
                    f"{len(pending_group_calls)} 个分组 = {total_calls} 次")

        # Episode-level disputes: each result is a dict of field→decision
        # We use _run_judge_tasks with a custom store_fn via a wrapper
        ep_lock = threading.Lock()

        def _judge_ep(ep_key: str, kwargs: dict) -> tuple[str, dict]:
            result = judge_episode_batch(**kwargs, pool=pool, model=model)
            # Expand field-level results into cache keys (thread-safe)
            with ep_lock:
                for f, decision in result.items():
                    judge_cache[f"{ep_key}||{f}"] = decision
            return ep_key, result

        # Source-line group disputes
        def _judge_grp(cache_key: str, kwargs: dict) -> tuple[str, dict]:
            decision = judge_source_line_group(**kwargs, pool=pool, model=model)
            return cache_key, decision

        # Run group calls
        _run_judge_tasks(
            pending_group_calls, _judge_grp, judge_cache,
            max_workers, step_label="judge step1 group", model=model,
        )
        # Run episode calls; results are expanded inside _judge_ep via closure,
        # so we pass a dummy cache that we discard.
        _ep_dummy: dict = {}
        _run_judge_tasks(
            pending_episode_calls, _judge_ep, _ep_dummy,
            max_workers, step_label="judge step1 ep", model=model,
        )

    # Save judge cache
    judge_cache_path.write_text(
        json.dumps(judge_cache, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # Build and save merged_episodes.json
    llm1_cache = load_json_cache(logs_dir / "llm1_step1_results.json")
    llm2_cache = load_json_cache(logs_dir / "llm2_step1_results.json")

    merged_all: dict[str, dict] = {}
    for name, ds_item in llm1_cache.items():
        vf_item = llm2_cache.get(name, {})
        episodes = build_merged_episodes(name, ds_item, vf_item, judge_cache)
        merged_all[name] = {
            "episodes": episodes,
            "_meta": {"name": name, "source": "merged_step1"},
        }

    merged_path = logs_dir / "merged_episodes.json"
    save_json_cache(merged_path, merged_all)
    logger.info(f"Step1 Judge 完成: {len(judge_cache)} 裁决, {len(merged_all)} 人 merged_episodes")

    return merged_path


# ── Judge Step 2: rank disputes ─────────────────────────────────────────────

def judge_step2(
    logs_dir: Path,
    force: bool = False,
    max_workers: int = DEFAULT_WORKERS,
    pool: RoundRobinClientPool | None = None,
    model: str = "",
) -> Path:
    """Judge step2 rank disputes. Saves step2_judge_decisions.json."""
    logger.info("=== Step2 Judge: rank 裁判 ===")

    diff_path = logs_dir / "step2_diff_report.json"
    diff_report = json.loads(diff_path.read_text(encoding="utf-8"))

    judge_cache_path = logs_dir / "step2_judge_decisions.json"
    judge_cache: dict = {}
    if judge_cache_path.exists() and not force:
        try:
            judge_cache = json.loads(judge_cache_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    pending_calls: list[tuple[str, dict]] = []

    for person in diff_report:
        name = person["official_name"]
        for diff in person.get("diffs", []):
            if diff.get("scope") != "rank":
                continue
            ep_idx = diff["episode_idx"]
            cache_key = _rank_cache_key(name, ep_idx)
            if cache_key not in judge_cache:
                pending_calls.append((cache_key, {
                    "name": name,
                    "episode_idx": ep_idx,
                    "unit": diff.get("供职单位", ""),
                    "position": diff.get("职务", ""),
                    "llm1_rank": diff["llm1_value"],
                    "llm2_rank": diff["llm2_value"],
                }))

    if pending_calls:
        logger.info(f"  裁判调用: {len(pending_calls)} 个 rank 争议")

        def _judge_rank(cache_key: str, kwargs: dict) -> tuple[str, dict]:
            prompt = (
                f"官员：{kwargs['name']}\n"
                f"经历#{kwargs['episode_idx']}: {kwargs['unit']} {kwargs['position']}\n\n"
                f"LLM1判断级别：{kwargs['llm1_rank']}\n"
                f"LLM2判断级别：{kwargs['llm2_rank']}\n\n"
                "请判断哪个更准确。"
            )
            decision = _call_judge(_judge_system_rank(), prompt, pool=pool, model=model)
            return cache_key, decision

        _run_judge_tasks(
            pending_calls, _judge_rank, judge_cache,
            max_workers, step_label="judge step2", model=model,
        )

    judge_cache_path.write_text(
        json.dumps(judge_cache, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    logger.info(f"Step2 Judge 完成: {len(judge_cache)} 裁决")
    return judge_cache_path


# ── Judge Step 3: label disputes ────────────────────────────────────────────

def judge_step3(
    logs_dir: Path,
    force: bool = False,
    max_workers: int = DEFAULT_WORKERS,
    pool: RoundRobinClientPool | None = None,
    model: str = "",
) -> Path:
    """Judge step3 label/bio disputes. Saves step3_judge_decisions.json."""
    logger.info("=== Step3 Judge: label + bio 裁判 ===")

    diff_path = logs_dir / "step3_diff_report.json"
    diff_report = json.loads(diff_path.read_text(encoding="utf-8"))

    judge_cache_path = logs_dir / "step3_judge_decisions.json"
    judge_cache: dict = {}
    if judge_cache_path.exists() and not force:
        try:
            judge_cache = json.loads(judge_cache_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    pending_calls: list[tuple[str, dict]] = []

    for person in diff_report:
        name = person["official_name"]
        for diff in person.get("diffs", []):
            scope = diff.get("scope", "")
            field = diff.get("field", "")
            cache_key = _label_cache_key(name, field)
            if cache_key not in judge_cache:
                if scope == "label":
                    pending_calls.append((cache_key, {
                        "name": name, "field": field, "scope": scope,
                        "llm1_value": str(diff.get("llm1_value", "")),
                        "llm1_reason": diff.get("ds_reason", ""),
                        "llm2_value": str(diff.get("llm2_value", "")),
                        "llm2_reason": diff.get("qw_reason", ""),
                        "original_text_snippet": "",
                    }))
                elif scope in ("bio", "corruption"):
                    pending_calls.append((cache_key, {
                        "name": name, "field": field, "scope": scope,
                        "llm1_value": str(diff.get("llm1_value", "")),
                        "llm1_reason": "",
                        "llm2_value": str(diff.get("llm2_value", "")),
                        "llm2_reason": "",
                        "original_text_snippet": "",
                    }))

    if pending_calls:
        logger.info(f"  裁判调用: {len(pending_calls)} 个 label/bio 争议")

        def _judge_lbl(cache_key: str, kwargs: dict) -> tuple[str, dict]:
            decision = get_judge_decision(**kwargs, pool=pool, model=model)
            return cache_key, decision

        _run_judge_tasks(
            pending_calls, _judge_lbl, judge_cache,
            max_workers, step_label="judge step3", model=model,
        )

    judge_cache_path.write_text(
        json.dumps(judge_cache, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    logger.info(f"Step3 Judge 完成: {len(judge_cache)} 裁决")
    return judge_cache_path


# ── Excel styling helpers ────────────────────────────────────────────────────

FILL_RED = PatternFill(fill_type="solid", fgColor="FFCCCC")
FILL_GREY = PatternFill(fill_type="solid", fgColor="E0E0E0")
FILL_BLUE = PatternFill(fill_type="solid", fgColor="CCE5FF")
FILL_PURPLE = PatternFill(fill_type="solid", fgColor="E8CCFF")
FILL_GREEN = PatternFill(fill_type="solid", fgColor="CCFFCC")
HEADER_FONT = Font(bold=True)


def _set_header_colors(ws, col_groups: dict):
    fills = {"DS": FILL_GREY, "LLM1": FILL_GREY, "QW": FILL_BLUE, "LLM2": FILL_BLUE, "裁判": FILL_PURPLE, "Final": FILL_GREEN}
    for group, cols in col_groups.items():
        fill = fills.get(group)
        if not fill:
            continue
        for col_idx in cols:
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = fill
            cell.font = HEADER_FONT


def _auto_width(ws, cap=45):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=6)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, cap)
