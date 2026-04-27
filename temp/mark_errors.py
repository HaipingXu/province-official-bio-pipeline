import json, copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict

# Load errors
with open("temp/all_errors.json") as f:
    errors = json.load(f)

# Load workbook
wb = load_workbook("output/浙江/浙江_officials.xlsx")
ws = wb.active

# Build header map: col_name -> col_index (1-based)
headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        headers[val.strip()] = col

print(f"Headers: {list(headers.keys())}")
print(f"Total rows: {ws.max_row}, Total cols: {ws.max_column}")

# Build official -> list of row indices in xlsx
# Column "姓名" tells us the official name; "经历序号" tells us the row number within that official
name_col = headers.get("姓名")
seq_col = headers.get("经历序号")
print(f"姓名 col: {name_col}, 经历序号 col: {seq_col}")

# Map (name, seq) -> xlsx_row
cell_map = {}
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=name_col).value
    seq = ws.cell(row=r, column=seq_col).value
    if name and seq is not None:
        cell_map[(str(name).strip(), int(seq))] = r

print(f"Mapped {len(cell_map)} (name, seq) -> xlsx_row entries")

# Apply blue fill
blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

marked = 0
missed = 0
col_errors = defaultdict(int)  # column -> error count
col_totals = defaultdict(int)  # column -> total cells

# Count totals per column
for col_name, col_idx in headers.items():
    col_totals[col_name] = ws.max_row - 1  # exclude header

for err in errors:
    name = err["name"]
    row = err["row"]
    col_name = err["column"]
    
    key = (name, row)
    xlsx_row = cell_map.get(key)
    col_idx = headers.get(col_name)
    
    if xlsx_row and col_idx:
        ws.cell(row=xlsx_row, column=col_idx).fill = blue_fill
        col_errors[col_name] += 1
        marked += 1
    else:
        missed += 1
        if not xlsx_row:
            pass  # print(f"  Missing row: {key}")
        if not col_idx:
            print(f"  Missing col: '{col_name}'")

print(f"\nMarked {marked} cells blue, {missed} missed")

# Save
wb.save("output/浙江/浙江_officials.xlsx")
print("Saved xlsx")

# Error rate table
print("\n" + "="*60)
print(f"{'Column':<25} {'Errors':>7} {'Total':>7} {'Rate':>8}")
print("="*60)
for col_name in sorted(col_errors.keys(), key=lambda x: -col_errors[x]):
    total = col_totals[col_name]
    errs = col_errors[col_name]
    rate = errs / total * 100 if total > 0 else 0
    print(f"{col_name:<25} {errs:>7} {total:>7} {rate:>7.1f}%")

total_cells = sum(col_totals.values())
total_errs = sum(col_errors.values())
print("-"*60)
print(f"{'TOTAL':<25} {total_errs:>7} {total_cells:>7} {total_errs/total_cells*100:>7.2f}%")
