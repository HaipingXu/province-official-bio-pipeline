"""
gold.py — Parse ground truth from test5_officials.xlsx

Returns structured gold data for Mode 1 (step1+4) and Mode 2 (step2+3) evaluation.

Ground truth schema (per official):
  person_fields: 升迁_省长, 升迁_省委书记, 本省提拔, 本省学习, 是否落马
  episodes: list of dicts with step1+2+3 fields
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# ── Paths ────────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).parent.parent
GOLD_XLSX = PROJECT_ROOT / "output" / "test5" / "test5_officials.xlsx"
PREPROCESSED_JSON = PROJECT_ROOT / "logs" / "test5" / "preprocessed_texts.json"
MERGED_EPISODES_JSON = PROJECT_ROOT / "logs" / "test5" / "merged_episodes.json"

# ── Column name aliases ──────────────────────────────────────────────────────

# Person-level fields we care about
PERSON_FIELDS = [
    "升迁_省长", "升迁_省委书记", "本省提拔", "本省学习",
    "出生年份", "籍贯", "籍贯（市）", "少数民族", "女性", "全日制本科",
    "是否落马", "落马原因",
]

# Episode-level fields by step
STEP1_FIELDS = ["起始时间", "终止时间", "供职单位", "职务"]
STEP2_FIELDS = ["组织标签", "标志位", "任职地（省）", "任职地（市）", "中央/地方"]
STEP3_FIELDS = ["该条行政级别"]

ALL_EPISODE_FIELDS = STEP1_FIELDS + STEP2_FIELDS + STEP3_FIELDS


def _normalize_val(v: Any) -> Any:
    """Normalize cell values: strip whitespace, convert None/''/nan to None."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "None", "nan", "NaN"):
        return None
    return s


def load_gold(xlsx_path: Path = GOLD_XLSX) -> dict[str, dict]:
    """
    Load test5_officials.xlsx and return gold data keyed by official name.

    Return structure:
    {
        "习近平": {
            "person": {
                "升迁_省长": 0,
                "升迁_省委书记": 0,
                "本省提拔": 0,
                "本省学习": 0,
                ...
            },
            "episodes": [
                {
                    "episode_idx": 1,
                    "起始时间": "1969.00",
                    "终止时间": "1975.00",
                    "供职单位": "...",
                    "职务": "...",
                    "组织标签": "...",
                    "标志位": "...",
                    "任职地（省）": "...",
                    "任职地（市）": "...",
                    "中央/地方": "...",
                    "该条行政级别": "...",
                }
            ]
        }
    }
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    gold: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))

        name = str(row_dict.get("姓名", "")).strip()
        if not name:
            continue

        if name not in gold:
            gold[name] = {"person": {}, "episodes": []}
            # Extract person-level fields (same for all rows of this official)
            for f in PERSON_FIELDS:
                v = row_dict.get(f)
                if v is not None and str(v).strip() not in ("", "None", "nan", "NaN"):
                    try:
                        # Keep numeric fields as numbers
                        if f in ("升迁_省长", "升迁_省委书记", "本省提拔", "本省学习",
                                  "少数民族", "女性", "全日制本科", "出生年份"):
                            gold[name]["person"][f] = int(float(str(v)))
                        else:
                            gold[name]["person"][f] = str(v).strip()
                    except (ValueError, TypeError):
                        gold[name]["person"][f] = str(v).strip()

        # Extract episode-level fields
        ep: dict[str, Any] = {}

        # episode_idx from 经历序号
        ep_idx = row_dict.get("经历序号")
        try:
            ep["episode_idx"] = int(ep_idx)
        except (TypeError, ValueError):
            ep["episode_idx"] = len(gold[name]["episodes"]) + 1

        for f in ALL_EPISODE_FIELDS:
            v = row_dict.get(f)
            ep[f] = _normalize_val(v)

        gold[name]["episodes"].append(ep)

    logger.info(f"[gold] Loaded {len(gold)} officials from {xlsx_path.name}: "
                f"{list(gold.keys())}")
    return gold


def load_preprocessed() -> dict[str, dict]:
    """Load preprocessed_texts.json, return {name: preprocessed_dict}."""
    data = json.loads(PREPROCESSED_JSON.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return {r["name"]: r for r in data if r and "name" in r}
    # Already a dict keyed by name
    return data


def load_merged_episodes() -> dict[str, list[dict]]:
    """
    Load merged_episodes.json (judge-decided step1 output).
    Returns {name: [episode_dict, ...]}
    """
    data = json.loads(MERGED_EPISODES_JSON.read_text(encoding="utf-8"))
    if isinstance(data, list):
        result: dict[str, list[dict]] = {}
        for r in data:
            if not isinstance(r, dict):
                continue
            name = r.get("_meta", {}).get("name", "")
            if name:
                result[name] = r.get("episodes", [])
        return result
    # Dict keyed by name
    return {
        name: rec.get("episodes", [])
        for name, rec in data.items()
    }


def get_step1_from_gold(gold: dict[str, dict]) -> dict[str, list[dict]]:
    """
    Extract step1 episodes from gold data.
    Used as input to Mode 2 (step2+3) evaluation.
    Returns {name: [{episode_idx, 供职单位, 职务, 起始时间, 终止时间}]}
    """
    result: dict[str, list[dict]] = {}
    for name, rec in gold.items():
        episodes = []
        for ep in rec["episodes"]:
            episodes.append({
                "episode_idx": ep["episode_idx"],
                "source_line": ep["episode_idx"],  # use episode_idx as proxy
                "起始时间": ep.get("起始时间") or "",
                "终止时间": ep.get("终止时间") or "",
                "供职单位": ep.get("供职单位") or "",
                "职务": ep.get("职务") or "",
            })
        result[name] = episodes
    return result


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    gold = load_gold()
    for name, rec in gold.items():
        print(f"\n{name}: {len(rec['episodes'])} episodes")
        print(f"  person: {rec['person']}")
        print(f"  episodes[0]: {rec['episodes'][0]}")
