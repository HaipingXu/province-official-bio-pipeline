"""
fix_gold.py — Fix known issues in test5_officials.xlsx before eval.

Auto-fixed (unambiguous):
  1. 组织标签 "国务院组成部门" → "国务院及其组成部门" (10 rows)
  2. 标志位 bracket '省委常委（其他)' → '省委常委（其他）' (1 row)

Flagged for user judgment (requires explicit --fix-city flag):
  3. 任职地（市）非空 for 省级标志位（省委书记/省委副书记等）rows (9 rows)
     These should be '' per step2_classify.md rule.

Usage:
  uv run eval/fix_gold.py --dry-run          # show changes, don't write
  uv run eval/fix_gold.py                     # write in-place (backup created)
  uv run eval/fix_gold.py --fix-city          # also fix 任职地（市）(needs user approval)
  uv run eval/fix_gold.py --out gold_fixed.xlsx  # write to different file
"""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

import openpyxl

GOLD_XLSX = Path(__file__).parent.parent / "output" / "test5" / "test5_officials.xlsx"

PROVINCE_FLAGS = {
    '省委书记', '省长', '省委副书记（省长）', '省委副书记（非省长）',
    '副省长（常委）', '副省长（非常委）', '省常委（其他）', '省组织部长', '省组织部副部长',
}

# Only organizations based in Beijing with no sub-national location
# EXCLUDES: 教育部直属高校/部属高校（物理上有校园）, 科研院所（中央）, 直属机构/部委管理的国家局
#   (国家局可能在外地驻点，不一定在北京)
CENTRAL_BEIJING_ORG_TAGS = {
    '党中央机关',
    '国务院及其组成部门',
    '全国人大机关',
    '全国政协机关',
    '最高法院/最高检察院',
    '中央军委机关',
    '共青团中央',
    '民主党派中央机关',
    '全国性人民团体',
}


def fix_gold(
    src: Path = GOLD_XLSX,
    dst: Path | None = None,
    dry_run: bool = False,
    fix_city: bool = False,
) -> list[str]:
    """
    Apply fixes to gold xlsx.
    Returns list of change descriptions.
    """
    changes: list[str] = []
    wb = openpyxl.load_workbook(src)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    def col(name: str) -> int:
        return headers.index(name) + 1

    org_col = col('组织标签')
    flag_col = col('标志位')
    city_col = col('任职地（市）')
    prov_col = col('任职地（省）')
    name_col = col('姓名')
    ep_col = col('经历序号')

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        get = lambda c: str(row[c-1].value or '').strip()
        set_ = lambda c, v: None if dry_run else setattr(row[c-1], 'value', v)

        name = get(name_col)
        ep = get(ep_col)
        org  = get(org_col)
        flag = get(flag_col)
        city = get(city_col)
        prov = get(prov_col)

        # Fix 1: 组织标签 旧名
        if org == '国务院组成部门':
            desc = f"  {name} ep#{ep}: 组织标签 '国务院组成部门' → '国务院及其组成部门'"
            changes.append(desc)
            set_(org_col, '国务院及其组成部门')

        # Fix 2: 标志位 bracket mismatch
        if flag == '省委常委（其他)':
            desc = f"  {name} ep#{ep}: 标志位 '省委常委（其他)' → '省委常委（其他）'"
            changes.append(desc)
            set_(flag_col, '省委常委（其他）')

        cl = get(col('中央/地方'))

        # Fix 3: 任职地（市）for province-level flags (requires --fix-city)
        # Rule: 省级单位 → 任职地（市）必须留空
        if fix_city and flag in PROVINCE_FLAGS and city:
            desc = f"  {name} ep#{ep}: 任职地（市）'{city}' → '' [flag={flag}]"
            changes.append(desc)
            set_(city_col, '')

        # Fix 3b: 任职地（市）for central Beijing-based organizations
        # Rule: 中央机关（北京总部）→ 任职地（市）必须留空
        # Uses CENTRAL_BEIJING_ORG_TAGS to avoid clearing city for universities,
        # research institutes, and regional outposts of national bureaus.
        elif fix_city and org in CENTRAL_BEIJING_ORG_TAGS and city:
            desc = f"  {name} ep#{ep}: 任职地（市）'{city}' → '' [org={org}]"
            changes.append(desc)
            set_(city_col, '')

    if not dry_run:
        out = dst or src
        if out == src:
            # Backup original
            bak = src.with_suffix('.xlsx.bak')
            shutil.copy2(src, bak)
            print(f"[backup] {bak}")
        wb.save(out)
        print(f"[saved] {out}")

    return changes


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true',
                        help='Show changes without writing')
    parser.add_argument('--fix-city', action='store_true',
                        help='Also clear 任职地（市）for provincial/central posts')
    parser.add_argument('--out', type=str, default='',
                        help='Output path (default: in-place with backup)')
    args = parser.parse_args()

    dst = Path(args.out) if args.out else None
    mode = "DRY RUN" if args.dry_run else "APPLY"
    print(f"[{mode}] fix_gold.py {'--fix-city' if args.fix_city else ''}")
    print()

    changes = fix_gold(
        dry_run=args.dry_run,
        fix_city=args.fix_city,
        dst=dst,
    )

    if changes:
        print(f"{'Changes (not applied):' if args.dry_run else 'Applied changes:'}")
        for c in changes:
            print(c)
        print(f"\nTotal: {len(changes)} changes")
    else:
        print("No changes needed.")


if __name__ == '__main__':
    main()
