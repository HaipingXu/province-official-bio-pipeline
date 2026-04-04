"""
Phase 5 (v2): Excel Export

Produces three output files:
  output/{city}_officials.xlsx    — all rows for the city (A→AC, 29 columns)
  output/{city}_mayors.xlsx       — rows where 该条是省长 == 1
  output/{city}_secretaries.xlsx  — rows where 该条是省委书记 == 1

Additionally the battle.xlsx produced by battle_generator.py is kept in output/.

Column order follows config.COLUMNS exactly.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import COLUMNS, LOGS_DIR, OUTPUT_DIR

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Style constants ────────────────────────────────────────────────────────────

FILL_HEADER    = PatternFill(fill_type="solid", fgColor="2F5496")   # dark blue
FILL_MAYOR     = PatternFill(fill_type="solid", fgColor="E2EFDA")   # light green
FILL_SECRETARY = PatternFill(fill_type="solid", fgColor="DAEEF3")   # light blue
FILL_ALT       = PatternFill(fill_type="solid", fgColor="F9F9F9")   # off-white alt row
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=10)
NORMAL_FONT    = Font(size=10)
CENTER_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

FILL_RED_CONF  = PatternFill(fill_type="solid", fgColor="FFCCCC")   # red for low confidence

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# Columns that look better centred
CENTRE_COLS = {
    "年份", "出生年份", "少数民族", "女性", "全日制本科",
    "升迁_省长", "升迁_省委书记", "本省提拔", "本省学习",
    "是否当过省长", "是否当过省委书记", "最终行政级别",
    "经历序号", "该条行政级别", "中央/地方", "是否落马", "该条是省长", "该条是省委书记",
}

# Fixed column widths (characters)
COL_WIDTHS: dict[str, int] = {
    "年份": 8, "省份": 10, "城市": 10, "姓名": 8,
    "出生年份": 10, "籍贯": 12, "籍贯（市）": 14,
    "少数民族": 10, "女性": 6,
    "全日制本科": 10, "升迁_省长": 12, "升迁_省委书记": 14,
    "本省提拔": 10, "本省学习": 10,
    "是否当过省长": 14, "是否当过省委书记": 16, "最终行政级别": 14,
    "经历序号": 8, "起始时间": 10, "终止时间": 10,
    "组织标签": 18, "标志位": 20, "该条行政级别": 14,
    "供职单位": 30, "职务": 22,
    "原文引用": 40,
    "争议未解决": 35,
    "裁判理由": 40,
    "任职地（省）": 16, "任职地（市）": 14, "中央/地方": 10,
    "": 4,  # spacer
    "是否落马": 10, "落马原因": 30, "备注栏": 28,
    "该条是省长": 12, "该条是省委书记": 14,
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _rows_to_df(rows: list[dict]) -> pd.DataFrame:
    """Convert list of row dicts → DataFrame with COLUMNS column order."""
    if not rows:
        return pd.DataFrame(columns=COLUMNS)
    df = pd.DataFrame(rows)
    # Ensure all COLUMNS present (fill missing with "")
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[COLUMNS]


def _apply_styles(ws, df: pd.DataFrame, highlight_col: str | None = None):
    """Apply header styles, column widths, alternating rows, and optional highlight."""
    headers = list(df.columns)
    n_rows = len(df)

    # Header row
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill  = FILL_HEADER
        cell.font  = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        # Column width
        w = COL_WIDTHS.get(col_name, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Row height for header
    ws.row_dimensions[1].height = 20

    # Determine highlight fill for each data row
    highlight_col_idx = (headers.index(highlight_col) + 1) if highlight_col and highlight_col in headers else None

    for row_idx in range(2, n_rows + 2):
        # Base alternating fill
        base_fill = FILL_ALT if row_idx % 2 == 0 else None

        # Check if this row should be highlighted (mayor/secretary row)
        is_highlighted = False
        if highlight_col_idx:
            val = ws.cell(row=row_idx, column=highlight_col_idx).value
            is_highlighted = val == 1

        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font   = NORMAL_FONT
            cell.border = THIN_BORDER

            if col_name in CENTRE_COLS:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

            if is_highlighted:
                # Mayor row → green; secretary row → blue (determined by highlight_col)
                if highlight_col == "该条是省长":
                    cell.fill = FILL_MAYOR
                elif highlight_col == "该条是省委书记":
                    cell.fill = FILL_SECRETARY
            elif base_fill:
                cell.fill = base_fill

        ws.row_dimensions[row_idx].height = 15

    # Highlight dispute cells with low confidence (contains [信心:<90])
    dispute_col_idx = (headers.index("争议未解决") + 1) if "争议未解决" in headers else None
    if dispute_col_idx:
        for row_idx in range(2, n_rows + 2):
            cell = ws.cell(row=row_idx, column=dispute_col_idx)
            val = str(cell.value or "")
            if "[信心:" in val:
                import re as _re
                for m in _re.finditer(r'\[信心:(\d+)\]', val):
                    if int(m.group(1)) < 90:
                        cell.fill = FILL_RED_CONF
                        break

    ws.freeze_panes = "E2"   # freeze person-level cols + header


def write_excel(df: pd.DataFrame, path: Path, sheet_name: str = "数据",
                highlight_col: str | None = None):
    """Write a single-sheet Excel file with full styling."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        _apply_styles(ws, df, highlight_col=highlight_col)
    print(f"  ✓ 已保存: {path.name}  ({len(df)} 行)")


# ── Main export ────────────────────────────────────────────────────────────────

def run_export(
    city: str,
    province: str,
    final_rows_path: Path | None = None,
    output_dir: Path | None = None,
) -> dict[str, Path]:
    """
    Load final_rows.json, write three Excel files.
    Returns dict of {label: Path}.
    """
    print(f"\n=== Phase 5 (v2): Excel 导出 ===")

    if final_rows_path is None:
        final_rows_path = LOGS_DIR / "final_rows.json"
    if output_dir is None:
        output_dir = OUTPUT_DIR

    output_dir.mkdir(parents=True, exist_ok=True)

    if not final_rows_path.exists():
        print(f"  ✗ final_rows.json 不存在: {final_rows_path}")
        return {}

    rows: list[dict] = json.loads(final_rows_path.read_text(encoding="utf-8"))
    df_all = _rows_to_df(rows)

    # Detect province mode: use governor/secretary labels for file names and sheets
    _province_names = {
        "北京", "天津", "河北", "山西", "内蒙古", "辽宁", "吉林", "黑龙江",
        "上海", "江苏", "浙江", "安徽", "福建", "江西", "山东",
        "河南", "湖北", "湖南", "广东", "广西", "海南",
        "重庆", "四川", "贵州", "云南", "西藏",
        "陕西", "甘肃", "青海", "宁夏", "新疆",
    }
    _prov_mode = city in _province_names

    mayor_label = "省长"
    sec_label = "省委书记"
    mayors_file = "governors"

    # --- All officials ---
    path_all = output_dir / f"{city}_officials.xlsx"
    write_excel(df_all, path_all, sheet_name="全部履历")

    # --- Governors: ALL rows for anyone who was ever a governor ---
    mayor_names = df_all[df_all["是否当过省长"] == 1]["姓名"].unique()
    df_mayors = df_all[df_all["姓名"].isin(mayor_names)].reset_index(drop=True)
    path_mayors = output_dir / f"{city}_{mayors_file}.xlsx"
    write_excel(df_mayors, path_mayors, sheet_name=f"{mayor_label}履历", highlight_col="该条是省长")

    # --- Secretaries: ALL rows for anyone who was ever a prov secretary ---
    sec_names = df_all[df_all["是否当过省委书记"] == 1]["姓名"].unique()
    df_secs = df_all[df_all["姓名"].isin(sec_names)].reset_index(drop=True)
    path_secs = output_dir / f"{city}_secretaries.xlsx"
    write_excel(df_secs, path_secs, sheet_name=f"{sec_label}履历", highlight_col="该条是省委书记")

    print(f"\n  全部: {len(df_all)} 行")
    print(f"  {mayor_label}: {len(df_mayors)} 行 ({df_mayors['姓名'].nunique() if not df_mayors.empty else 0} 人)")
    print(f"  {sec_label}: {len(df_secs)} 行 ({df_secs['姓名'].nunique() if not df_secs.empty else 0} 人)")

    return {
        "all":        path_all,
        "mayors":     path_mayors,
        "secretaries": path_secs,
    }
