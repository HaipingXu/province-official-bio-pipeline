"""
Phase 5 (v2): Excel Export

Produces three output files:
  output/{city}_officials.xlsx    — all rows for the city (A→AD, 29 columns)
  output/{city}_governors.xlsx    — all rows for people who ever served as governor
  output/{city}_secretaries.xlsx  — all rows for people who ever served as prov secretary

Additionally the battle.xlsx produced by judge.py is kept in output/.

Column order follows config.COLUMNS exactly.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import COLUMNS, DATA_DIR, LOGS_DIR, OUTPUT_DIR, PROVINCE_NAMES

# ── Style constants ────────────────────────────────────────────────────────────

FILL_HEADER    = PatternFill(fill_type="solid", fgColor="2F5496")   # dark blue
FILL_MAYOR     = PatternFill(fill_type="solid", fgColor="E2EFDA")   # light green
FILL_SECRETARY = PatternFill(fill_type="solid", fgColor="DAEEF3")   # light blue
FILL_ALT       = PatternFill(fill_type="solid", fgColor="F9F9F9")   # off-white alt row
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=10)
NORMAL_FONT    = Font(size=10)
CENTER_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

FILL_RED_CONF    = PatternFill(fill_type="solid", fgColor="FFCCCC")   # red for low confidence
FILL_RED_BLOCKED = PatternFill(fill_type="solid", fgColor="FF4444")   # deep red for blocked judge

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# Columns that look better centred
CENTRE_COLS = {
    "年份", "出生年份", "少数民族", "女性", "全日制本科",
    "升迁_省长", "升迁_省委书记", "本省提拔", "本省学习",
    "最终行政级别",
    "经历序号", "该条行政级别", "中央/地方", "是否落马",
}

# Fixed column widths (characters)
COL_WIDTHS: dict[str, int] = {
    "年份": 8, "省份": 10, "姓名": 8,
    "出生年份": 10, "籍贯": 12, "籍贯（市）": 14,
    "少数民族": 10, "女性": 6,
    "全日制本科": 10, "升迁_省长": 12, "升迁_省委书记": 14,
    "本省提拔": 10, "本省学习": 10, "最终行政级别": 14,
    "经历序号": 8, "起始时间": 10, "终止时间": 10,
    "组织标签": 18, "标志位": 20, "该条行政级别": 14,
    "供职单位": 30, "职务": 22,
    "原文引用": 40,
    "争议未解决": 35,
    "任职地（省）": 16, "任职地（市）": 14, "中央/地方": 10,
    "是否落马": 10, "落马原因": 30, "备注栏": 28,
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _rows_to_df(rows: list[dict]) -> pd.DataFrame:
    """Convert list of row dicts → DataFrame with COLUMNS column order."""
    if not rows:
        return pd.DataFrame(columns=COLUMNS)
    df = pd.DataFrame(rows)
    # Ensure all COLUMNS present (fill missing with "")
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[COLUMNS]


def _apply_styles(
    ws,
    df: pd.DataFrame,
    highlight_col: str | None = None,
    highlight_vals: set | None = None,
    highlight_fill: "PatternFill | None" = None,
):
    """Apply header styles, column widths, alternating rows, and optional highlight."""
    headers = list(df.columns)
    n_rows = len(df)

    # Header row
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill  = FILL_HEADER
        cell.font  = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        # Column width
        w = COL_WIDTHS.get(col_name, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Row height for header
    ws.row_dimensions[1].height = 20

    # Determine highlight fill for each data row
    highlight_col_idx = (headers.index(highlight_col) + 1) if highlight_col and highlight_col in headers else None

    for row_idx in range(2, n_rows + 2):
        # Base alternating fill
        base_fill = FILL_ALT if row_idx % 2 == 0 else None

        # Check if this row should be highlighted
        is_highlighted = False
        if highlight_col_idx:
            val = ws.cell(row=row_idx, column=highlight_col_idx).value
            if highlight_vals is not None:
                is_highlighted = val in highlight_vals
            else:
                is_highlighted = bool(val)

        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font   = NORMAL_FONT
            cell.border = THIN_BORDER

            if col_name in CENTRE_COLS:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

            if is_highlighted and highlight_fill:
                cell.fill = highlight_fill
            elif base_fill:
                cell.fill = base_fill

        ws.row_dimensions[row_idx].height = 15

    # Highlight dispute cells: deep red for blocked judge, light red for low confidence
    dispute_col_idx = (headers.index("争议未解决") + 1) if "争议未解决" in headers else None
    if dispute_col_idx:
        import re as _re
        for row_idx in range(2, n_rows + 2):
            cell = ws.cell(row=row_idx, column=dispute_col_idx)
            val = str(cell.value or "")
            if "[裁判被拦截]" in val:
                cell.fill = FILL_RED_BLOCKED
            elif "[信心:" in val:
                for m in _re.finditer(r'\[信心:(\d+)\]', val):
                    if int(m.group(1)) < 90:
                        cell.fill = FILL_RED_CONF
                        break

    ws.freeze_panes = "E2"   # freeze person-level cols + header


def write_excel(
    df: pd.DataFrame,
    path: Path,
    sheet_name: str = "数据",
    highlight_col: str | None = None,
    highlight_vals: set | None = None,
    highlight_fill: "PatternFill | None" = None,
):
    """Write a single-sheet Excel file with full styling."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        _apply_styles(ws, df, highlight_col=highlight_col,
                      highlight_vals=highlight_vals, highlight_fill=highlight_fill)
    print(f"  ✓ 已保存: {path.name}  ({len(df)} 行)")


# ── Main export ────────────────────────────────────────────────────────────────

def _load_names_from_txt(txt_path: Path) -> tuple[set[str], set[str]]:
    """
    Parse an officials.txt file and return (governor_names, secretary_names).
    Uses input_parser_province to keep parsing logic in one place.
    Returns (set of governor names, set of secretary names).
    """
    from input_parser_province import parse_province_officials_txt

    # Derive province short name from filename (e.g. "浙江_officials.txt" → "浙江")
    province_short = txt_path.stem.replace("_officials", "")
    parsed = parse_province_officials_txt(province_short, data_dir=txt_path.parent)

    governor_names = {o["name"] for o in parsed.get("governors", [])}
    secretary_names = {o["name"] for o in parsed.get("secretaries", [])}
    return governor_names, secretary_names


def run_export(
    province: str,
    final_rows_path: Path | None = None,
    output_dir: Path | None = None,
    officials_txt_path: Path | None = None,
) -> dict[str, Path]:
    """
    Load final_rows.json, write three Excel files.

    Args:
        officials_txt_path: Path to the officials.txt name list (e.g.
            data/1990/浙江_officials.txt). When provided, governor and
            secretary sheets are filtered by the names listed there.
            When None, falls back to inferring names from the 标志位 column.

    Returns dict of {label: Path}.
    """
    print(f"\n=== Phase 5 (v2): Excel 导出 ===")

    if final_rows_path is None:
        final_rows_path = LOGS_DIR / "final_rows.json"
    if output_dir is None:
        output_dir = OUTPUT_DIR

    output_dir.mkdir(parents=True, exist_ok=True)

    if not final_rows_path.exists():
        print(f"  ✗ final_rows.json 不存在: {final_rows_path}")
        return {}

    rows: list[dict] = json.loads(final_rows_path.read_text(encoding="utf-8"))
    df_all = _rows_to_df(rows)

    # --- All officials ---
    path_all = output_dir / f"{province}_officials.xlsx"
    write_excel(df_all, path_all, sheet_name="全部履历")

    # Governor / secretary highlight tags (for row colouring only)
    GOV_TAGS = {"省长", "省委副书记（省长）"}
    SEC_TAGS = {"省委书记"}

    # --- Determine governor / secretary name sets ---
    if officials_txt_path is not None and officials_txt_path.exists():
        governor_names, secretary_names = _load_names_from_txt(officials_txt_path)
        print(f"  从名单文件读取: 省长 {len(governor_names)} 人, 省委书记 {len(secretary_names)} 人")
    else:
        # Fallback: infer from 标志位 column
        governor_names = set(df_all[df_all["标志位"].isin(GOV_TAGS)]["姓名"].unique())
        secretary_names = set(df_all[df_all["标志位"].isin(SEC_TAGS)]["姓名"].unique())
        if officials_txt_path is not None:
            print(f"  ⚠ 名单文件不存在 ({officials_txt_path}), 回退到标志位推断")

    # --- Governors: ALL rows for anyone listed as governor in the txt ---
    df_mayors = df_all[df_all["姓名"].isin(governor_names)].reset_index(drop=True)
    path_mayors = output_dir / f"{province}_governors.xlsx"
    write_excel(df_mayors, path_mayors, sheet_name="省长履历",
                highlight_col="标志位", highlight_vals=GOV_TAGS, highlight_fill=FILL_MAYOR)

    # --- Secretaries: ALL rows for anyone listed as secretary in the txt ---
    df_secs = df_all[df_all["姓名"].isin(secretary_names)].reset_index(drop=True)
    path_secs = output_dir / f"{province}_secretaries.xlsx"
    write_excel(df_secs, path_secs, sheet_name="省委书记履历",
                highlight_col="标志位", highlight_vals=SEC_TAGS, highlight_fill=FILL_SECRETARY)

    print(f"\n  全部: {len(df_all)} 行")
    print(f"  省长: {len(df_mayors)} 行 ({df_mayors['姓名'].nunique() if not df_mayors.empty else 0} 人)")
    print(f"  省委书记: {len(df_secs)} 行 ({df_secs['姓名'].nunique() if not df_secs.empty else 0} 人)")

    return {
        "all":        path_all,
        "mayors":     path_mayors,
        "secretaries": path_secs,
    }
