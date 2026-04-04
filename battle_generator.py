"""
Phase 3b (v2): Battle Table Generator + Kimi K2.5 Judge

For each disputed field (DS ≠ QW), calls Kimi K2.5 as judge.
Fallback: DeepSeek if Kimi unavailable.

Outputs:
  output/{city}_battle.xlsx   — full parallel comparison with red highlights
  logs/judge_decisions.json   — judge verdicts + reasoning
"""

import json
import logging
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pandas as pd
from openai import OpenAI
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from config import (
    LOGS_DIR, OUTPUT_DIR,
    KIMI_API_KEY, KIMI_API_KEYS, KIMI_BASE_URL, KIMI_MODEL,
    DEEPSEEK_API_KEY, DEEPSEEK_API_KEYS, DEEPSEEK_BASE_URL, DEEPSEEK_MODEL,
    DEFAULT_WORKERS,
    KIMI_RPM_LIMIT, KIMI_TPM_LIMIT,
)
from utils import extract_json, RoundRobinClientPool, SmoothRateLimiter


def _to_float_date(s: str) -> float:
    """Parse YYYY.MM to float for comparison."""
    try:
        parts = str(s).split(".")
        yr = int(parts[0])
        mo = int(parts[1]) if len(parts) > 1 and parts[1] not in ("00", "") else 0
        return yr + mo / 12.0
    except Exception:
        return -1.0

logger = logging.getLogger(__name__)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Module-level judge client pool (created once, reused) ─────────────────────

_judge_pool: RoundRobinClientPool | None = None
_judge_model: str = ""
_fallback_pool: RoundRobinClientPool | None = None
_init_lock = threading.Lock()
_judge_rate_limiter: SmoothRateLimiter | None = None


def _init_judge_clients() -> None:
    """Lazily initialise judge client pools on first use (thread-safe)."""
    global _judge_pool, _judge_model, _fallback_pool

    if _judge_pool is not None:
        return

    with _init_lock:
        # Double-check after acquiring lock
        if _judge_pool is not None:
            return

        # Primary: Kimi K2.5 (multi-key pool)
        kimi_keys = KIMI_API_KEYS if KIMI_API_KEYS else ([KIMI_API_KEY] if KIMI_API_KEY else [])
        ds_keys = DEEPSEEK_API_KEYS if DEEPSEEK_API_KEYS else ([DEEPSEEK_API_KEY] if DEEPSEEK_API_KEY else [])

        if kimi_keys:
            _judge_pool = RoundRobinClientPool(kimi_keys, KIMI_BASE_URL)
            _judge_model = KIMI_MODEL
            logger.info(f"裁判模型: Kimi K2.5 ({KIMI_MODEL}), {_judge_pool.size} keys")
            global _judge_rate_limiter
            _judge_rate_limiter = SmoothRateLimiter(rpm_limit=KIMI_RPM_LIMIT, tpm_limit=KIMI_TPM_LIMIT)
            logger.info(f"  速率限制: RPM={KIMI_RPM_LIMIT}, TPM={KIMI_TPM_LIMIT:,}")
        elif ds_keys:
            _judge_pool = RoundRobinClientPool(ds_keys, DEEPSEEK_BASE_URL)
            _judge_model = DEEPSEEK_MODEL
            logger.warning("KIMI_API_KEY 未设置，使用 DeepSeek 作为裁判")
        else:
            logger.error("无可用裁判 API key（KIMI / DEEPSEEK）")

        # DeepSeek R1 (reasoning model) as fallback if Kimi is primary
        if kimi_keys and ds_keys:
            _fallback_pool = RoundRobinClientPool(ds_keys, DEEPSEEK_BASE_URL)


# ── Judge decision ────────────────────────────────────────────────────────────

# System prompt: per-episode batch judge with "自行修正" capability
JUDGE_SYSTEM = (
    "你是一名中国政治数据核查专家，根据原文对两个LLM提取结果做裁判。\n\n"
    "重要规则：\n"
    "1. 学习经历（本科、研究生、博士、进修、培训等）也是条目。\n"
    "2. 如果两方都不完全正确，你可以给出自己根据原文判断的正确值（verdict=自行修正）。\n"
    "3. 党委系统供职单位必须使用全称带「中共」前缀（如「中共深圳市委」而非「深圳市委」、「中共广东省委」而非「广东省委」）。"
    "如果一方写了「中共X委」而另一方省略了「中共」，应采纳带「中共」的版本或自行修正为全称。\n"
    "4. 除党委命名规则外，其他字段忠实于百度百科原文用词。\n\n"
    "对每个争议字段，输出以下格式的JSON：\n"
    "{\"verdict\": \"采纳DS\"|\"采纳QW\"|\"自行修正\"|\"两者均存疑\", "
    "\"correct_value\": \"仅当verdict=自行修正时填写正确值\", "
    "\"confidence\": 0-100, "
    "\"reason\": \"<50字理由>\"}\n\n"
    "confidence 表示你对该裁决的信心程度（0=完全不确定，100=完全确定）。\n\n"
    "如果一次裁判多个字段，输出JSON对象，key为字段名，value为上述格式：\n"
    "{\"字段A\": {\"verdict\": ..., \"confidence\": 85, \"reason\": ...}, \"字段B\": {...}}\n\n"
    "只输出JSON，无任何其他文字。"
)

# Legacy single-field system prompt (for label judging)
JUDGE_SYSTEM_LABEL = (
    "你是一名中国政治数据核查专家，根据原文对两个LLM提取结果做裁判。\n"
    "党委系统供职单位必须使用全称带「中共」前缀（如「中共深圳市委」而非「深圳市委」）。\n"
    "你必须且只能输出一个JSON对象，格式为：\n"
    "{\"verdict\": \"采纳DS\"|\"采纳QW\"|\"自行修正\"|\"两者均存疑\", "
    "\"correct_value\": \"仅当verdict=自行修正时填写\", "
    "\"confidence\": 0-100, "
    "\"reason\": \"<50字理由>\"}\n"
    "confidence 表示你对该裁决的信心程度（0=完全不确定，100=完全确定）。\n"
    "不要输出任何其他文字、解释或代码块标记。"
)


def _call_judge(system: str, prompt: str) -> dict:
    """
    Call judge (Kimi K2.5 primary, DeepSeek fallback).
    Returns parsed JSON dict.
    """
    _init_judge_clients()

    # Try primary judge (with retry for Kimi's temperature=1 randomness)
    if _judge_pool:
        is_kimi = "kimi" in _judge_model.lower()
        judge_temp = 1.0 if is_kimi else 0.0
        max_retries = 3 if is_kimi else 1
        for attempt in range(max_retries):
            try:
                if _judge_rate_limiter:
                    _judge_rate_limiter.acquire(estimated_tokens=2000)
                client = _judge_pool.next_client()
                # Kimi K2.5 thinking model: no seed, larger max_tokens
                extra_kwargs = {} if is_kimi else {"seed": 42}
                resp = client.chat.completions.create(
                    model=_judge_model,
                    messages=[
                        {"role": "system", "content": system},
                        {"role": "user", "content": prompt},
                    ],
                    temperature=judge_temp,
                    max_tokens=16384,
                    **extra_kwargs,
                )
                content = resp.choices[0].message.content or ""
                if not content.strip():
                    logger.warning(f"[{_judge_model}] empty response (attempt {attempt+1})")
                    continue
                result = extract_json(content)
                result["judge_model"] = _judge_model
                return result
            except Exception as e:
                err_str = str(e)
                logger.warning(f"[{_judge_model} judge error] {err_str}")
                # Content filter / high risk → skip retries, fall through to fallback
                if "content_filter" in err_str or "high risk" in err_str or "high_risk" in err_str:
                    logger.warning(f"[{_judge_model}] content filtered, falling through to fallback")
                    break
                if attempt < max_retries - 1:
                    continue

    # Try fallback: DeepSeek R1 reasoning model
    _FALLBACK_MODEL = "deepseek-reasoner"
    if _fallback_pool:
        try:
            client = _fallback_pool.next_client()
            resp = client.chat.completions.create(
                model=_FALLBACK_MODEL,
                messages=[
                    {"role": "user", "content": system + "\n\n" + prompt},
                ],
                max_tokens=8192,
            )
            result = extract_json(resp.choices[0].message.content)
            result["judge_model"] = f"deepseek_fallback/{_FALLBACK_MODEL}"
            return result
        except Exception as e:
            logger.error(f"[DeepSeek fallback error] {e}")

    return {"verdict": "两者均存疑", "reason": "裁判调用失败", "judge_model": "error"}


def judge_source_line_group(
    name: str,
    line_num: int,
    raw_text: str,
    ds_episodes: list[dict],
    vf_episodes: list[dict],
) -> dict:
    """
    Judge a source_line where DS and VF extracted different numbers of episodes.
    The judge sees ALL episodes (all fields) from both sides, decides which split
    to adopt, AND outputs the final corrected episodes with all fields.
    Returns {"adopt": "DS"/"VF", "reason": "...", "episodes": [...], "judge_model": "..."}
    """
    all_fields = [
        "起始时间", "终止时间", "组织标签", "标志位", "供职单位", "职务",
        "任职地（省）", "任职地（市）", "中央/地方",
    ]

    def _fmt(eps):
        lines = []
        for i, ep in enumerate(eps, 1):
            parts = [f"{f}={ep.get(f, '')}" for f in all_fields]
            lines.append(f"  #{i}: " + ", ".join(parts))
        return "\n".join(lines) if lines else "  （无）"

    prompt = (
        f"官员：{name}\n"
        f"原文行 L{line_num:02d}: {raw_text}\n\n"
        f"DS从此行提取了 {len(ds_episodes)} 条经历：\n{_fmt(ds_episodes)}\n\n"
        f"VF从此行提取了 {len(vf_episodes)} 条经历：\n{_fmt(vf_episodes)}\n\n"
        "两个LLM对这一行原文的拆分方式不同。请判断哪种更准确，并输出最终正确版本。\n\n"
        "判断规则：\n"
        "- 同一时间段在同一单位的多个职务 → 合并为一条（职务用顿号连接）\n"
        "- 不同时间段或不同单位 → 拆分为多条\n"
        "- 党组书记/党组成员与行政职务同期同单位 → 合并\n\n"
        "同时检查所有字段的准确性：\n"
        "- 组织标签：是否与供职单位匹配（如'市人民政府'应标'地级市政府'）\n"
        "- 供职单位：是否含完整层级（如学院名应含大学名前缀）\n"
        "- 中央/地方：中央部委/直属机构/全国人大政协='中央'，省市县='地方'\n"
        "- 任职地：省份用全称（如'广东省'非'广东'），直辖市省市均填\n\n"
        f"输出JSON:\n"
        f"{{\n"
        f"  \"adopt\": \"DS\" 或 \"VF\",\n"
        f"  \"confidence\": 0-100,\n"
        f"  \"reason\": \"<50字理由>\",\n"
        f"  \"episodes\": [\n"
        f"    {{\n"
        f"      \"source_line\": {line_num},\n"
        f"      \"起始时间\": \"YYYY.MM\",\n"
        f"      \"终止时间\": \"YYYY.MM\",\n"
        f"      \"组织标签\": \"...\",\n"
        f"      \"供职单位\": \"...\",\n"
        f"      \"职务\": \"...\",\n"
        f"      \"任职地（省）\": \"...\",\n"
        f"      \"任职地（市）\": \"...\",\n"
        f"      \"中央/地方\": \"中央或地方\"\n"
        f"    }}\n"
        f"  ]\n"
        f"}}\n"
        f"confidence 表示你对该裁决的信心程度（0=完全不确定，100=完全确定）。\n"
        f"episodes数组包含最终正确的经历条目（可基于adopt方修正细节）。\n"
        f"只输出JSON，无任何其他文字。"
    )

    return _call_judge(JUDGE_SYSTEM, prompt)


def judge_episode_batch(
    name: str,
    disputed_fields: list[str],
    ds_values: dict[str, str],
    qw_values: dict[str, str],
    ref_ds: str,
    ref_qw: str,
) -> dict[str, dict]:
    """
    Judge ALL disputed fields for one episode in a single API call.
    Returns {field_name: {"verdict": ..., "reason": ..., "correct_value": ...}}
    """
    if not disputed_fields:
        return {}

    # Build per-field comparison block
    field_lines = []
    for f in disputed_fields:
        field_lines.append(f"  {f}: DS=「{ds_values.get(f, '')}」 QW=「{qw_values.get(f, '')}」")
    fields_block = "\n".join(field_lines)

    prompt = (
        f"官员：{name}\n\n"
        f"DS来源行：{ref_ds or '（无）'}\n"
        f"验证来源行：{ref_qw or '（无）'}\n\n"
        f"以下 {len(disputed_fields)} 个字段存在争议：\n{fields_block}\n\n"
        "请对每个字段分别裁判。\n"
        "输出JSON对象，key为字段名，value含verdict/reason/correct_value。"
    )

    result = _call_judge(JUDGE_SYSTEM, prompt)

    # Parse result — it might be a flat dict (single field) or nested dict (multi-field)
    parsed: dict[str, dict] = {}

    if len(disputed_fields) == 1:
        # Single field — result IS the verdict dict
        f = disputed_fields[0]
        if "verdict" in result:
            parsed[f] = result
        elif f in result:
            parsed[f] = result[f]
        else:
            parsed[f] = {"verdict": "两者均存疑", "reason": "解析失败"}
    else:
        # Multi-field — result should be {field: {verdict, reason}}
        for f in disputed_fields:
            if f in result and isinstance(result[f], dict):
                parsed[f] = result[f]
            elif "verdict" in result:
                # Fallback: judge returned single verdict for all fields
                parsed[f] = result
            else:
                parsed[f] = {"verdict": "两者均存疑", "reason": "批量裁判未覆盖此字段"}

    # Propagate judge_model to all sub-results
    model = result.get("judge_model", _judge_model)
    for f in parsed:
        parsed[f]["judge_model"] = model

    return parsed


def get_judge_decision(
    name: str,
    field: str,
    scope: str,
    ds_value: str,
    ds_reason: str,
    qw_value: str,
    qw_reason: str,
    original_text_snippet: str,
) -> dict:
    """
    Call judge for ONE disputed field (label scope).
    Returns {"verdict": ..., "reason": ..., "correct_value": ...}
    """
    prompt = (
        f"官员：{name}  字段：{field}（{scope}）\n\n"
        f"原文依据：{original_text_snippet or '（无）'}\n\n"
        f"DeepSeek提取：{ds_value}\n"
        f"DS依据：{ds_reason or '（无）'}\n\n"
        f"Qwen提取：{qw_value}\n"
        f"QW依据：{qw_reason or '（无）'}\n\n"
        "根据原文和两方依据，判断哪个更准确。如果两方都不正确，请自行给出正确值。"
    )

    return _call_judge(JUDGE_SYSTEM_LABEL, prompt)


# ── Build battle rows ──────────────────────────────────────────────────────────

def _build_rank_map(step3_data: dict) -> dict[int, str]:
    """Build episode_idx → final_rank map from step3 result."""
    ranks = step3_data.get("ranks", [])
    return {r.get("episode_idx", 0): r.get("final_rank", "") for r in ranks}


def build_battle_rows(diff_report: list[dict], judge_cache: dict,
                      max_workers: int = DEFAULT_WORKERS,
                      career_lines_by_name: dict[str, dict[int, str]] | None = None,
                      ) -> tuple[list[dict], list[dict]]:
    """
    Returns:
      episode_rows: one row per matched episode pair
      label_rows: one row per label field
    """
    from verifier_v2 import group_by_source_line

    if career_lines_by_name is None:
        career_lines_by_name = {}

    episode_rows: list[dict] = []
    label_rows: list[dict] = []

    # Track source_lines with count mismatches for group judging
    # (name, line_num) -> {"ds": [...], "vf": [...], "raw": "..."}
    sl_mismatch_data: dict[tuple[str, int], dict] = {}

    for person in diff_report:
        name = person["official_name"]
        ds_s1 = person.get("ds_step1", {})
        vf_s1 = person.get("vf_step1", person.get("qw_step1", {}))
        ds_s2 = person.get("ds_step2", {})
        vf_s2 = person.get("vf_step2", person.get("qw_step2", {}))
        ds_s3 = person.get("ds_step3", {})
        vf_s3 = person.get("vf_step3", {})

        eps_ds = ds_s1.get("episodes", [])
        eps_vf = vf_s1.get("episodes", [])

        # Rank maps: episode_idx → final_rank
        ds_rank_map = _build_rank_map(ds_s3)
        vf_rank_map = _build_rank_map(vf_s3)

        # Career lines map for this official (source_line → raw_text)
        cl_map = career_lines_by_name.get(name, {})

        # Group by source_line for matching
        ds_groups = group_by_source_line(eps_ds)
        vf_groups = group_by_source_line(eps_vf)
        all_lines = sorted(set(ds_groups) | set(vf_groups))

        for line_num in all_lines:
            ds_list = ds_groups.get(line_num, [])
            vf_list = vf_groups.get(line_num, [])

            # Track count mismatch for group judging
            if len(ds_list) != len(vf_list):
                sl_mismatch_data[(name, line_num)] = {
                    "ds_episodes": ds_list,
                    "vf_episodes": vf_list,
                    "raw_text": cl_map.get(line_num, ""),
                }

            # Sort both by unit name for stable comparison
            ds_sorted = sorted(ds_list, key=lambda e: e.get("供职单位", ""))
            vf_sorted = sorted(vf_list, key=lambda e: e.get("供职单位", ""))

            max_len = max(len(ds_sorted), len(vf_sorted))
            for i in range(max_len):
                ep_ds = ds_sorted[i] if i < len(ds_sorted) else None
                ep_vf = vf_sorted[i] if i < len(vf_sorted) else None

                row: dict = {"官员姓名": name, "source_line": line_num}

                ep_fields = ["起始时间", "终止时间", "组织标签", "标志位",
                             "供职单位", "职务", "行政级别",
                             "任职地（省）", "任职地（市）", "中央/地方"]

                if ep_ds:
                    sl_num = ep_ds.get('source_line', line_num)
                    raw = cl_map.get(sl_num, "")
                    row["source_line_DS"] = f"L{sl_num:02d}: {raw}" if raw else f"L{sl_num:02d}"
                    ep_idx = ep_ds.get("经历序号", i + 1)
                    for f in ep_fields:
                        if f == "行政级别":
                            row[f"DS_{f}"] = ds_rank_map.get(ep_idx, "")
                        else:
                            row[f"DS_{f}"] = ep_ds.get(f, "")
                else:
                    row["source_line_DS"] = ""
                    for f in ep_fields:
                        row[f"DS_{f}"] = "【DS无此条】"

                if ep_vf:
                    sl_num = ep_vf.get('source_line', line_num)
                    raw = cl_map.get(sl_num, "")
                    row["source_line_VF"] = f"L{sl_num:02d}: {raw}" if raw else f"L{sl_num:02d}"
                    ep_idx = ep_vf.get("经历序号", i + 1)
                    for f in ep_fields:
                        if f == "行政级别":
                            row[f"QW_{f}"] = vf_rank_map.get(ep_idx, "")
                        else:
                            row[f"QW_{f}"] = ep_vf.get(f, "")
                else:
                    row["source_line_VF"] = ""
                    for f in ep_fields:
                        row[f"QW_{f}"] = "【VF无此条】"

                # Find all diffs
                check_fields = ["起始时间", "终止时间", "组织标签", "标志位", "供职单位", "职务",
                                "行政级别", "任职地（省）", "任职地（市）", "中央/地方"]
                disputed_fields = []
                for f in check_fields:
                    v_ds = str(row.get(f"DS_{f}", ""))
                    v_vf = str(row.get(f"QW_{f}", ""))
                    if v_ds != v_vf and v_ds != "【DS无此条】" and v_vf != "【VF无此条】":
                        disputed_fields.append(f)

                is_mismatch = len(ds_list) != len(vf_list)
                row["存在差异"] = "YES" if (disputed_fields or is_mismatch) else "NO"
                row["差异字段"] = ", ".join(disputed_fields) if disputed_fields else ("数量不匹配" if is_mismatch else "")
                row["_disputed_fields"] = disputed_fields
                row["_is_mismatch_group"] = is_mismatch
                episode_rows.append(row)

        # ─ Label comparison rows ───────────────────────────────────────────────
        for field in ["升迁_省长", "升迁_省委书记", "本省提拔", "本省学习"]:
            v_ds = ds_s2.get(field)
            v_qw = vf_s2.get(field)
            row = {
                "官员姓名": name, "字段": field,
                "DS值": v_ds, "DS依据": ds_s2.get(field + "依据", ""),
                "QW值": v_qw, "QW依据": vf_s2.get(field + "依据", ""),
                "存在差异": "YES" if v_ds != v_qw else "NO",
                "_needs_judge": v_ds != v_qw and v_ds is not None and v_qw is not None,
            }
            label_rows.append(row)

    # ── Batch judge calls (per-episode, concurrent) ───────────────────────────
    # Collect pending judge requests: episode-level (batch per episode) + label-level
    pending_episode_calls: list[tuple[str, dict]] = []  # (ep_cache_key, call_kwargs)
    pending_label_calls: list[tuple[str, dict]] = []

    for row in episode_rows:
        name = row["官员姓名"]
        disputed = row.get("_disputed_fields", [])

        # Skip per-episode judging for count-mismatch source_lines
        # (handled by group judge below)
        if row.get("_is_mismatch_group"):
            continue

        if not disputed:
            continue
        # Episode-level cache key (per episode, not per field)
        ep_key = f"{name}||ep_batch||{row.get('DS_供职单位', '')}||{row.get('DS_职务', '')}"
        # Check if ALL fields for this episode are already cached
        all_cached = all(
            f"{ep_key}||{f}" in judge_cache for f in disputed
        )
        if not all_cached:
            ds_vals = {f: str(row.get(f"DS_{f}", "")) for f in disputed}
            qw_vals = {f: str(row.get(f"QW_{f}", "")) for f in disputed}
            pending_episode_calls.append((ep_key, {
                "name": name,
                "disputed_fields": disputed,
                "ds_values": ds_vals,
                "qw_values": qw_vals,
                "ref_ds": row.get("source_line_DS", ""),
                "ref_qw": row.get("source_line_VF", ""),
            }))
            # Store ep_key on the row for later lookup
            row["_ep_cache_key"] = ep_key

    # Group judge calls for source_lines with count mismatches
    pending_group_calls: list[tuple[str, dict]] = []
    for (name, ln), data in sl_mismatch_data.items():
        cache_key = f"{name}||sl_group||{ln}"
        if cache_key not in judge_cache:
            pending_group_calls.append((cache_key, {
                "name": name,
                "line_num": ln,
                "raw_text": data["raw_text"],
                "ds_episodes": data["ds_episodes"],
                "vf_episodes": data["vf_episodes"],
            }))

    for row in label_rows:
        if row.get("_needs_judge"):
            name = row["官员姓名"]
            field = row["字段"]
            cache_key = f"{name}||label||{field}"
            if cache_key not in judge_cache:
                pending_label_calls.append((cache_key, {
                    "name": name, "field": field, "scope": "label",
                    "ds_value": str(row["DS值"]),
                    "ds_reason": row.get("DS依据", ""),
                    "qw_value": str(row["QW值"]),
                    "qw_reason": row.get("QW依据", ""),
                    "original_text_snippet": "",
                }))

    total_calls = len(pending_episode_calls) + len(pending_group_calls) + len(pending_label_calls)

    # Execute pending calls concurrently
    if total_calls > 0:
        lock = threading.Lock()
        n_fields = sum(len(kw["disputed_fields"]) for _, kw in pending_episode_calls)
        print(f"  裁判调用: {len(pending_episode_calls)} 条履历({n_fields}个字段) + "
              f"{len(pending_group_calls)} 个分组 + "
              f"{len(pending_label_calls)} 个标签 = {total_calls} 次API调用 (workers={max_workers})")

        def _judge_episode(ep_key: str, kwargs: dict) -> tuple[str, dict[str, dict]]:
            """Judge all disputed fields for one episode in a single call."""
            result = judge_episode_batch(**kwargs)
            return ep_key, result

        def _judge_label(cache_key: str, kwargs: dict) -> tuple[str, dict]:
            decision = get_judge_decision(**kwargs)
            return cache_key, decision

        def _judge_group(cache_key: str, kwargs: dict) -> tuple[str, dict]:
            decision = judge_source_line_group(**kwargs)
            return cache_key, decision

        if max_workers <= 1:
            for ep_key, kwargs in pending_episode_calls:
                _, field_results = _judge_episode(ep_key, kwargs)
                for f, decision in field_results.items():
                    judge_cache[f"{ep_key}||{f}"] = decision
                time.sleep(0.5)
            for cache_key, kwargs in pending_group_calls:
                _, decision = _judge_group(cache_key, kwargs)
                judge_cache[cache_key] = decision
                time.sleep(0.5)
            for cache_key, kwargs in pending_label_calls:
                _, decision = _judge_label(cache_key, kwargs)
                judge_cache[cache_key] = decision
                time.sleep(0.5)
        else:
            with ThreadPoolExecutor(max_workers=max_workers) as pool:
                # Submit episode batch calls
                ep_futures = {
                    pool.submit(_judge_episode, ek, kw): ek
                    for ek, kw in pending_episode_calls
                }
                # Submit group calls
                grp_futures = {
                    pool.submit(_judge_group, ck, kw): ck
                    for ck, kw in pending_group_calls
                }
                # Submit label calls
                lb_futures = {
                    pool.submit(_judge_label, ck, kw): ck
                    for ck, kw in pending_label_calls
                }

                for fut in as_completed({**ep_futures, **grp_futures, **lb_futures}):
                    try:
                        if fut in ep_futures:
                            ep_key, field_results = fut.result()
                            with lock:
                                for f, decision in field_results.items():
                                    judge_cache[f"{ep_key}||{f}"] = decision
                        elif fut in grp_futures:
                            ck, decision = fut.result()
                            with lock:
                                judge_cache[ck] = decision
                        else:
                            ck, decision = fut.result()
                            with lock:
                                judge_cache[ck] = decision
                    except Exception as e:
                        ck = ep_futures.get(fut) or grp_futures.get(fut) or lb_futures.get(fut, "?")
                        logger.error(f"[judge error] {ck}: {e}")
                        with lock:
                            if fut in ep_futures:
                                # Mark all fields as disputed
                                kwargs = dict(pending_episode_calls)[ck] if isinstance(ck, str) else {}
                                for f in kwargs.get("disputed_fields", []):
                                    judge_cache[f"{ck}||{f}"] = {
                                        "verdict": "两者均存疑", "reason": f"异常: {e}", "judge_model": "error"
                                    }
                            else:
                                judge_cache[ck] = {
                                    "verdict": "两者均存疑", "reason": f"异常: {e}", "judge_model": "error"
                                }

    # ── Populate judge results into rows ──────────────────────────────────────
    check_fields = ["起始时间", "终止时间", "组织标签", "标志位", "供职单位", "职务",
                    "行政级别", "任职地（省）", "任职地（市）", "中央/地方"]

    for row in episode_rows:
        name = row["官员姓名"]
        disputed_fields = row.pop("_disputed_fields", [])
        is_mismatch = row.pop("_is_mismatch_group", False)
        ep_key = row.pop("_ep_cache_key", f"{name}||ep_batch||{row.get('DS_供职单位', '')}||{row.get('DS_职务', '')}")
        judge_verdicts, judge_reasons, final_vals = [], [], {}

        if is_mismatch:
            # Group verdict for count-mismatch source_lines
            ln = row.get("source_line", 0)
            grp_key = f"{name}||sl_group||{ln}"
            grp_decision = judge_cache.get(grp_key, {"adopt": "DS", "reason": "未裁判"})
            adopt = grp_decision.get("adopt", "DS")
            reason = grp_decision.get("reason", "")
            confidence = grp_decision.get("confidence", "")
            row["裁判结论"] = f"整行采纳{adopt}"
            row["裁判理由"] = reason
            row["裁判信心"] = confidence

            # Final values use the adopted side
            for f in check_fields:
                if adopt == "VF":
                    row[f"Final_{f}"] = row.get(f"QW_{f}", row.get(f"DS_{f}", ""))
                else:
                    row[f"Final_{f}"] = row.get(f"DS_{f}", "")
        else:
            confidences = []
            for f in disputed_fields:
                cache_key = f"{ep_key}||{f}"
                decision = judge_cache.get(cache_key, {"verdict": "两者均存疑", "reason": "未裁判"})
                verdict = decision.get("verdict", "")
                conf = decision.get("confidence", "")
                judge_verdicts.append(f"{f}→{verdict}")
                judge_reasons.append(f"{f}：{decision.get('reason', '')}")
                if conf != "":
                    confidences.append(conf)

                if verdict == "采纳DS":
                    final_vals[f] = row.get(f"DS_{f}", "")
                elif verdict == "采纳QW":
                    final_vals[f] = row.get(f"QW_{f}", "")
                elif verdict == "自行修正":
                    corrected = decision.get("correct_value", "")
                    if corrected:
                        final_vals[f] = corrected
                    else:
                        final_vals[f] = row.get(f"DS_{f}", "")  # fallback to DS
                else:
                    final_vals[f] = f"[争议] DS={row.get(f'DS_{f}', '')} QW={row.get(f'QW_{f}', '')}"

            row["裁判结论"] = " | ".join(judge_verdicts) if judge_verdicts else "无争议"
            row["裁判理由"] = " | ".join(judge_reasons) if judge_reasons else ""
            # Min confidence across all disputed fields
            row["裁判信心"] = min(confidences) if confidences else ""

            for f in check_fields:
                row[f"Final_{f}"] = final_vals.get(f, row.get(f"DS_{f}", ""))

    for row in label_rows:
        needs_judge = row.pop("_needs_judge", False)
        if needs_judge:
            name = row["官员姓名"]
            field = row["字段"]
            cache_key = f"{name}||label||{field}"
            decision = judge_cache.get(cache_key, {"verdict": "两者均存疑", "reason": "未裁判"})
            verdict = decision.get("verdict", "")
            row["裁判结论"] = verdict
            row["裁判理由"] = decision.get("reason", "")
            row["裁判信心"] = decision.get("confidence", "")
            if verdict == "采纳DS":
                row["最终值"] = row["DS值"]
            elif verdict == "采纳QW":
                row["最终值"] = row["QW值"]
            elif verdict == "自行修正":
                corrected = decision.get("correct_value", "")
                row["最终值"] = corrected if corrected else row["DS值"]
            else:
                row["最终值"] = "[争议]"
        else:
            row["裁判结论"] = "无争议"
            row["裁判理由"] = ""
            row["裁判信心"] = ""
            row["最终值"] = row["DS值"]

    return episode_rows, label_rows


# ── Excel styling ──────────────────────────────────────────────────────────────

FILL_RED = PatternFill(fill_type="solid", fgColor="FFCCCC")
FILL_GREY = PatternFill(fill_type="solid", fgColor="E0E0E0")
FILL_BLUE = PatternFill(fill_type="solid", fgColor="CCE5FF")
FILL_PURPLE = PatternFill(fill_type="solid", fgColor="E8CCFF")
FILL_GREEN = PatternFill(fill_type="solid", fgColor="CCFFCC")

HEADER_FONT = Font(bold=True)


def _set_header_colors(ws, col_groups: dict):
    fills = {"DS": FILL_GREY, "QW": FILL_BLUE, "裁判": FILL_PURPLE, "Final": FILL_GREEN}
    for group, cols in col_groups.items():
        fill = fills.get(group)
        if not fill:
            continue
        for col_idx in cols:
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = fill
            cell.font = HEADER_FONT


def _auto_width(ws, cap=45):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=6)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, cap)


FILL_LOW_CONF = PatternFill(fill_type="solid", fgColor="FFE0CC")  # orange for low confidence


def save_battle_excel(episode_rows: list[dict], label_rows: list[dict],
                      path: Path, city: str):
    """Write battle.xlsx with two sheets and cell-level conditional formatting."""
    if not episode_rows and not label_rows:
        logger.warning("无对比数据，跳过 battle.xlsx 生成")
        return

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        if episode_rows:
            df_ep = pd.DataFrame(episode_rows)
            df_ep.to_excel(writer, sheet_name="履历对比", index=False)
            ws = writer.sheets["履历对比"]

            headers = list(df_ep.columns)
            ds_cols = [i + 1 for i, h in enumerate(headers) if h.startswith("DS_")]
            qw_cols = [i + 1 for i, h in enumerate(headers) if h.startswith("QW_")]
            judge_cols = [i + 1 for i, h in enumerate(headers) if h in ("裁判结论", "裁判理由", "裁判信心")]
            final_cols = [i + 1 for i, h in enumerate(headers) if h.startswith("Final_")]
            _set_header_colors(ws, {"DS": ds_cols, "QW": qw_cols, "裁判": judge_cols, "Final": final_cols})

            # Build header→col_idx mapping for field-name lookups
            header_idx = {h: i + 1 for i, h in enumerate(headers)}

            # Cell-level red: only highlight conflicting DS/QW cells, not entire row
            diff_fields_col = header_idx.get("差异字段", 0)
            conf_col = header_idx.get("裁判信心", 0)

            for row_idx in range(2, len(df_ep) + 2):
                # Parse disputed fields from the "差异字段" column
                diff_text = str(ws.cell(row=row_idx, column=diff_fields_col).value or "")
                if diff_text and diff_text != "NO" and diff_text != "":
                    disputed = [f.strip() for f in diff_text.split(",") if f.strip()]
                    for field_name in disputed:
                        if field_name == "数量不匹配":
                            # Mismatch group: highlight all DS and QW cells
                            for ci in ds_cols + qw_cols:
                                ws.cell(row=row_idx, column=ci).fill = FILL_RED
                            break
                        # Highlight DS_field and QW_field cells
                        ds_ci = header_idx.get(f"DS_{field_name}", 0)
                        qw_ci = header_idx.get(f"QW_{field_name}", 0)
                        if ds_ci:
                            ws.cell(row=row_idx, column=ds_ci).fill = FILL_RED
                        if qw_ci:
                            ws.cell(row=row_idx, column=qw_ci).fill = FILL_RED

                # Confidence < 90 → highlight confidence cell in orange
                if conf_col:
                    conf_val = ws.cell(row=row_idx, column=conf_col).value
                    try:
                        if conf_val is not None and int(conf_val) < 90:
                            ws.cell(row=row_idx, column=conf_col).fill = FILL_RED
                    except (ValueError, TypeError):
                        pass

            ws.freeze_panes = "A2"
            _auto_width(ws)

        if label_rows:
            df_lb = pd.DataFrame(label_rows)
            df_lb.to_excel(writer, sheet_name="标签对比", index=False)
            ws2 = writer.sheets["标签对比"]
            headers2 = list(df_lb.columns)
            header_idx2 = {h: i + 1 for i, h in enumerate(headers2)}
            diff_col = header_idx2.get("存在差异", 0)
            ds_col = header_idx2.get("DS值", 0)
            qw_col = header_idx2.get("QW值", 0)
            conf_col2 = header_idx2.get("裁判信心", 0)
            for row_idx in range(2, len(df_lb) + 2):
                if ws2.cell(row=row_idx, column=diff_col).value == "YES":
                    # Only highlight DS/QW value cells, not entire row
                    if ds_col:
                        ws2.cell(row=row_idx, column=ds_col).fill = FILL_RED
                    if qw_col:
                        ws2.cell(row=row_idx, column=qw_col).fill = FILL_RED
                # Confidence < 90
                if conf_col2:
                    conf_val = ws2.cell(row=row_idx, column=conf_col2).value
                    try:
                        if conf_val is not None and int(conf_val) < 90:
                            ws2.cell(row=row_idx, column=conf_col2).fill = FILL_RED
                    except (ValueError, TypeError):
                        pass
            ws2.freeze_panes = "A2"
            _auto_width(ws2)

    logger.info(f"Battle 表已保存: {path.name}")


# ── Main entry ─────────────────────────────────────────────────────────────────

def run_battle(
    diff_report_path: Path,
    output_dir: Path,
    city: str,
    force: bool = False,
    max_workers: int = DEFAULT_WORKERS,
    officials_dir: Path | None = None,
    logs_dir: Path | None = None,
) -> dict:
    logger.info("=== Phase 3b: Battle 表生成 + Kimi K2.5 裁判 ===")

    diff_report = json.loads(diff_report_path.read_text(encoding="utf-8"))

    # Build career_lines_by_name for all officials in the diff report
    from text_preprocessor import preprocess_official
    career_lines_by_name: dict[str, dict[int, str]] = {}
    for person in diff_report:
        oname = person.get("official_name", "")
        if oname:
            preprocessed = preprocess_official(oname, officials_dir=officials_dir)
            if preprocessed and preprocessed.get("career_lines"):
                career_lines_by_name[oname] = {
                    cl["line_num"]: cl["raw_text"] for cl in preprocessed["career_lines"]
                }

    # Load judge cache
    _logs = logs_dir or LOGS_DIR
    judge_cache_path = _logs / "judge_decisions.json"
    judge_cache: dict = {}
    if judge_cache_path.exists() and not force:
        try:
            judge_cache = json.loads(judge_cache_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    episode_rows, label_rows = build_battle_rows(
        diff_report, judge_cache, max_workers=max_workers,
        career_lines_by_name=career_lines_by_name,
    )

    # Save judge cache
    judge_cache_path.write_text(
        json.dumps(judge_cache, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    battle_path = output_dir / f"{city}_battle.xlsx"
    save_battle_excel(episode_rows, label_rows, battle_path, city)

    n_disputed = sum(1 for r in episode_rows if r.get("存在差异") == "YES")
    n_judged = sum(1 for r in episode_rows if r.get("存在差异") == "YES"
                   and r.get("裁判结论", "无争议") != "无争议")
    logger.info(f"履历行: {len(episode_rows)} 总计, {n_disputed} 存在差异, {n_judged} 已裁判")
    logger.info(f"标签行: {len(label_rows)} 总计")

    return {
        "battle_path": str(battle_path),
        "episode_rows": len(episode_rows),
        "label_rows": len(label_rows),
        "disputed": n_disputed,
    }
